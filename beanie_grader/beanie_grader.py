#!/usr/bin/env python3
"""
Beanie Size Grader
Automates size grading for knit beanies/hats with Excel output.

Author: YOUR NAME
"""

import os
import argparse
import pandas as pd
from openpyxl.styles import Font, PatternFill

# Default base specs and grading offsets
DEFAULT_BASE_SPECS = {
    'Crown Circumference (cm)': 56,
    'Height / Length (cm)': 22,
    'Brim Width (cm)': 8,
    'Gauge (stitches per 10cm)': 22
}

SIZES = ['XS', 'S', 'M', 'L', 'XL']
GRADE_OFFSETS = {
    'Crown Circumference (cm)': [-8, -4, 0, 4, 8],
    'Height / Length (cm)': [-2, -1, 0, 1, 2],
    'Brim Width (cm)': [0, 0, 0, 0.5, 1],
    'Gauge (stitches per 10cm)': [0, 0, 0, 0, 0]
}


def parse_args():
    parser = argparse.ArgumentParser(description="Knit Beanie Size Grader")
    parser.add_argument("--base_circumference", type=float, default=DEFAULT_BASE_SPECS['Crown Circumference (cm)'])
    parser.add_argument("--base_height", type=float, default=DEFAULT_BASE_SPECS['Height / Length (cm)'])
    parser.add_argument("--base_brim", type=float, default=DEFAULT_BASE_SPECS['Brim Width (cm)'])
    parser.add_argument("--base_gauge", type=int, default=DEFAULT_BASE_SPECS['Gauge (stitches per 10cm)'])
    parser.add_argument("--output", type=str, default="knit_beanie_graded_specs.xlsx")
    return parser.parse_args()


def build_graded_table(base_specs):
    df = pd.DataFrame({'Size': SIZES})
    for measurement, offsets in GRADE_OFFSETS.items():
        df[measurement] = [round(base_specs[measurement] + offset, 1) for offset in offsets]

    df['Est. Yarn Length (m) approx'] = (
        (df['Crown Circumference (cm)'] / 100 * 3.14 * df['Height / Length (cm)'] * 1.2 + 10) * 0.8 / 10
    ).round(1)

    df['Notes'] = ''
    df.loc[df['Crown Circumference (cm)'] < 50, 'Notes'] = 'Potentially too small'
    df.loc[df['Crown Circumference (cm)'] > 64, 'Notes'] = 'Potentially too large'

    return df


def write_to_excel(df, output_file):
    output_path = os.path.abspath(output_file)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Beanie Specs', index=False)
        ws = writer.sheets['Beanie Specs']

        for cell in ws[1]:
            cell.font = Font(bold=True)

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

        notes_col = df.columns.get_loc('Notes') + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=notes_col, max_col=notes_col):
            for cell in row:
                if cell.value == 'Potentially too small':
                    cell.fill = red_fill
                elif cell.value == 'Potentially too large':
                    cell.fill = yellow_fill

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 3

    print(f"\nExcel file created: {output_path}")


def main():
    args = parse_args()
    base_specs = {
        'Crown Circumference (cm)': args.base_circumference,
        'Height / Length (cm)': args.base_height,
        'Brim Width (cm)': args.base_brim,
        'Gauge (stitches per 10cm)': args.base_gauge
    }

    df = build_graded_table(base_specs)
    print("\nGenerated Graded Size Specs:")
    print(df)
    write_to_excel(df, args.output)


if __name__ == "__main__":
    main()
