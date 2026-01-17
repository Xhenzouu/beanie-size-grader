import os
import pandas as pd
from openpyxl.styles import Font, PatternFill
from beanie_grader.engine import build_graded_table
from beanie_grader.config import load_config
from beanie_grader.utils import validate_config

def run_command(args):
    """Handle the 'run' command: load config, build table, and export to Excel."""
    config = load_config(args.config)
    
    base_overrides = {}
    for name in config["measurements"]:
        arg_name = name.lower().replace(" ", "_").replace("/", "").replace("(", "").replace(")", "")
        base_overrides[name] = getattr(args, f"base_{arg_name}", None)
    
    df = build_graded_table(config, base_overrides)
    print("\nGenerated Graded Size Specs:")
    print(df)
    write_to_excel(df, args.output)

def validate_command(args):
    """Handle the 'validate' command: load and validate config."""
    config = load_config(args.config)
    valid = validate_config(config)
    if not valid:
        exit(1)

def list_command(args):
    """Handle the 'list' command: list available config files."""
    configs_dir = "configs"
    if not os.path.exists(configs_dir):
        print(f"No configs directory found at {configs_dir}.")
        return
    
    yaml_files = [f for f in os.listdir(configs_dir) if f.endswith('.yaml')]
    if not yaml_files:
        print("No YAML config files found in configs/.")
    else:
        print("Available config files:")
        for file in yaml_files:
            print(f"  - {file}")

def write_to_excel(df, output_file):
    """Helper function to write DataFrame to Excel with formatting."""
    output_path = os.path.abspath(output_file)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Product Specs', index=False)
        ws = writer.sheets['Product Specs']

        for cell in ws[1]:
            cell.font = Font(bold=True)

        if 'Notes' in df.columns:
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

            notes_col = df.columns.get_loc('Notes') + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=notes_col, max_col=notes_col):
                for cell in row:
                    if cell.value == 'Potentially too small':
                        cell.fill = red_fill
                    elif cell.value == 'Potentially too large':
                        cell.fill = yellow_fill

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    print(f"\nExcel file created: {output_path}")

def build_graded_table(config, base_overrides=None):
    """
    Build a graded table for product sizes, with product-specific enhancements if applicable.
    """
    base_overrides = base_overrides or {}
    sizes = config["sizes"]
    measurements = config["measurements"]
    
    df = pd.DataFrame({'Size': sizes})
    
    for measurement, spec in measurements.items():
        base_value = base_overrides.get(measurement, spec["base"])
        offsets = spec["offsets"]
        df[measurement] = [round(base_value + offset, 1) for offset in offsets]
    
    if config.get("name") == "beanie":
        if 'Crown Circumference (cm)' in df.columns and 'Height / Length (cm)' in df.columns:
            df['Est. Yarn Length (m) approx'] = (
                (df['Crown Circumference (cm)'] / 100 * 3.14 * df['Height / Length (cm)'] * 1.2 + 10) * 0.8 / 10
            ).round(1)
            
            df['Notes'] = ''
            df.loc[df['Crown Circumference (cm)'] < 50, 'Notes'] = 'Potentially too small'
            df.loc[df['Crown Circumference (cm)'] > 64, 'Notes'] = 'Potentially too large'
    
    return df