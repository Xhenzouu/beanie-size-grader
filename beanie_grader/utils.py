from openpyxl.styles import Font, PatternFill
import pandas as pd
from pathlib import Path

def validate_config(config: dict) -> bool:
    """
    Validate a product config YAML.

    Returns True if valid, False if errors found.
    """
    errors = []

    if "product" not in config or "name" not in config["product"]:
        errors.append("Missing product.name")

    sizes = config.get("sizes")
    if not sizes or not isinstance(sizes, list):
        errors.append("Missing or invalid sizes list")

    measurements = config.get("measurements")
    if not measurements or not isinstance(measurements, dict):
        errors.append("Missing or invalid measurements dict")
    else:
        for m_name, m_spec in measurements.items():
            if "base" not in m_spec:
                errors.append(f"Measurement '{m_name}' missing 'base'")
            if "offsets" not in m_spec or not isinstance(m_spec["offsets"], list):
                errors.append(f"Measurement '{m_name}' missing or invalid 'offsets'")
            elif len(m_spec["offsets"]) != len(sizes):
                errors.append(
                    f"Measurement '{m_name}' offsets length ({len(m_spec['offsets'])}) "
                    f"does not match sizes length ({len(sizes)})"
                )

    if errors:
        print("Config validation failed:")
        for e in errors:
            print(f"  - {e}")
        return False
    else:
        print(f"Config '{config['product']['name']}' is valid ✅")
        return True
    
def write_to_excel(df: pd.DataFrame, output_file: str):
    output_path = Path(output_file).resolve()

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Graded Specs', index=False)
        ws = writer.sheets['Graded Specs']

        for cell in ws[1]:
            cell.font = Font(bold=True)

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        notes_col = df.columns.get_loc('Notes') + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=notes_col, max_col=notes_col):
            for cell in row:
                if cell.value == 'Potentially too small':
                    cell.fill = red_fill
                elif cell.value == 'Potentially too large':
                    cell.fill = yellow_fill

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 3

        summary = df['Size'].value_counts().reindex(df['Size']).fillna(0).astype(int)
        summary_df = pd.DataFrame({'Size': summary.index, 'Count': summary.values})
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        ws_sum = writer.sheets['Summary']

        for cell in ws_sum[1]:
            cell.font = Font(bold=True)

        highlight_fill = PatternFill(start_color='FFEE99', end_color='FFEE99', fill_type='solid')
        for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value in ['XS', 'XL']:
                    cell.fill = highlight_fill

    print(f"\nExcel file created: {output_path}")